from openpyxl import Workbook
from openpyxl.styles import NamedStyle
from openpyxl.styles import PatternFill

# Imorting the necessary modules
try:
        from openpyxl.cell import get_column_letter
except ImportError:
        from openpyxl.utils import get_column_letter
        from openpyxl.utils import column_index_from_string

import os

class CSVCreator:
    def __init__(self, call_dict):
        self.call_dict = call_dict
        self.fileName = os.path.basename(os.getcwd()) + ".xlsx"
# "CallLogs1.xlsx"
        
    def createCSV(self):
        jitter_thrashhold = 15
        packloss_thrashold = 5.0
        delay_thrashold = 150
        workbook = Workbook()
        
        sheet = workbook.active
        sheet["A1"] = "No"
        sheet["B1"] = "MO"
        sheet["C1"] = "MT"
        sheet["D1"] = "Call Type"
        sheet["E1"] = "Call Initiate at"
        sheet["F1"] = "Call Starts at"
        sheet["G1"] = "Call Ends at"

        sheet["H1"] = "Call Direction "
        sheet["I1"] = "Network "
        sheet["J1"] = "Call Duration"
        sheet["K1"] = "Call Quality"
        sheet["L1"] = "Call Detail"
        sheet["M1"] = "Reason"

        sheet["N1"] = "Codec"
        sheet["O1"] = "AVG MOS"
        sheet["P1"] = "AVG Jitter"
        sheet["Q1"] = "AVG Latancy"
            
        sheet["R1"] = "MIN MOS"
        sheet["S1"] = "MIN Jitter"
        sheet["T1"] = "MIN Latancy"
            
        sheet["U1"] = "MAX MOS"
        sheet["V1"] = "MAX Jitter"
        sheet["W1"] = "MAX Latancy"

        
        sheet["X1"] = "WRG URL"
        sheet["Y1"] = "Call Id"
        
        # //self.call_type
        row = 2
        for callId, call in self.call_dict.items():
            # print(call.call_initiate_at)
            # print(call.call_id)
            suspiciousCall = None
            detailSheetId = call.call_id if len(call.call_id) <= 10 else callId[len(call.call_id)-10:]

            sheet.cell(column=1, row=row, value=row-1)
            sheet.cell(column=2, row=row, value=call.originatorAddress)
            sheet.cell(column=3, row=row, value=call.receiverAddress)
            sheet.cell(column=4, row=row, value=call.call_type)
            
            sheet.cell(column=5, row=row, value=call.call_initiate_at.strftime("%Y-%m-%d %H:%M:%S"))
            
    
            if call.call_ends_at is not None:
                sheet.cell(column=7, row=row, value=call.call_ends_at.strftime("%Y-%m-%d %H:%M:%S"))
                
            sheet.cell(column=8, row=row, value=call.callDirection)
            sheet.cell(column=9, row=row, value=call.networkTypes[0])
            sheet.cell(column=11, row=row, value=call.callQuality)
            if call.call_starts_at is not None and call.lastStatsDetails is not None:
                sheet.cell(column=6, row=row, value=call.call_starts_at.strftime("%Y-%m-%d %H:%M:%S"))
                sheet.cell(column=10, row=row, value=call.call_duration)
                if "audio_receive" in call.lastStatsDetails:
                    sheet.cell(column=14, row=row, value=call.lastStatsDetails["audio_receive"]["CodecName"])
                if call.lastMosDetails is not None and "calculated" in call.lastMosDetails:
                    sheet.cell(column=15, row=row, value=float(call.lastMosDetails["calculated"]["mos-avg"]))
                    sheet.cell(column=16, row=row, value=float(call.lastMosDetails["calculated"]["jitter-avg"]))
                    sheet.cell(column=17, row=row, value=float(call.lastMosDetails["calculated"]["latency-avg"]))
                    sheet.cell(column=18, row=row, value=float(call.lastMosDetails["calculated"]["mos-min"]))
                    sheet.cell(column=19, row=row, value=float(call.lastMosDetails["calculated"]["jitter-min"]))
                    sheet.cell(column=20, row=row, value=float(call.lastMosDetails["calculated"]["latency-min"]))
                    sheet.cell(column=21, row=row, value=float(call.lastMosDetails["calculated"]["mos-max"]))
                    sheet.cell(column=22, row=row, value=float(call.lastMosDetails["calculated"]["jitter-max"]))
                    sheet.cell(column=23, row=row, value=float(call.lastMosDetails["calculated"]["latency-max"]))
                else:
                    suspiciousCall = Suspicious("Mos Value is missing")
                    
            if call.incompleteCall is not None:
                print("else condition triggred for non start call")
                sheet.cell(column=12, row=row, value=call.incompleteCall.status)
                sheet.cell(column=13, row=row, value=call.incompleteCall.reason)
                calorCode = "55E2E9"
                if call.incompleteCall.status == "FAIL":
                    calorCode = "BC544B"
                for cell in sheet[row:row]:
                    cell.fill = PatternFill("solid", start_color=calorCode)
            elif suspiciousCall is not None:
                print("suspicious call idenfied")
                sheet.cell(column=12, row=row, value="SUSPICIOUS CALL")
                sheet.cell(column=13, row=row, value=suspiciousCall.reason)
                calorCode = "ffa500"
                for cell in sheet[row:row]:
                    cell.fill = PatternFill("solid", start_color=calorCode)
                
            
            sheet.cell(column=24, row=row, value=call.wrgUrl)
            link = "#'"+detailSheetId+"'!A1"
            sheet.cell(row=row, column=25).hyperlink = link
            sheet.cell(column=25, row=row, value=call.session_id)
            sheet.cell(row=row, column=25).style = "Hyperlink"
            
            
            # print(call.session_id)
            

            detailsSheet = workbook.create_sheet(detailSheetId)
                
            
            detailsSheet["A1"] = "No"
            detailsSheet["B1"] = "Time"
            detailsSheet["C1"] = "MOS"
            detailsSheet["D1"] = "Jitter"
            detailsSheet["E1"] = "Delay"
            detailsSheet["F1"] = "Packet Loss"
            detailsSheet["G1"] = "R"
            detailsSheet["H1"] = "Recieve Audio Packet Received"
            detailsSheet["I1"] = "Recieve Audio Packet lost"
            detailsSheet["J1"] = "Recieve Audio Packet SSRC"
            detailsSheet["K1"] = "Recieve Audio Packet JITTER RECEIVED"

            detailsSheet["L1"] = "Send Audio Packet Sent"
            detailsSheet["M1"] = "Send Audio Packet lost" 
            detailsSheet["N1"] = "Send Audio Packet SSRC"
            detailsSheet["O1"] = "Send Audio Packet Codec Bitrate" 
            detailsSheet["P1"] = "Send Audio Packet Jitter Received"
            detailsSheet["Q1"] = "Send Audio Packet RTT"
            detailsSheet["R1"] = "Remote Address"
            

            detailRow = 2
            for time, stat in call.stats.items():
                detailsSheet.cell(column=1, row=detailRow, value=detailRow-1)
                detailsSheet.cell(column=2, row=detailRow, value=time)
                # print(stat.isCallHold)

                mos = 0.0000
                jiiter = 0 
                delay = 0
                packet_loss = 0 
                r_value = 0 

                if stat.mos is not None and "pp_calculated" in stat.mos:
                    mos = float(stat.mos["pp_calculated"]["MOS"])
                    jiiter = int(stat.mos["pp_calculated"]["jitter"])
                    delay = int(stat.mos["pp_calculated"]["delay"])
                    packet_loss = float(stat.mos["pp_calculated"]["percentPacketLoss"])
                    r_value = int(stat.mos["pp_calculated"]["R"])

                recieveAudioPacketLost = 0
                recieveAudioPacketsReceived = 0 
                recieveAudioSsrc = ""
                recieveAudioJitterReceived = 0 
                
                if "audio_receive" in stat.stats:
                    recieveAudioPacketsReceived = int(stat.stats["audio_receive"].get("packetsReceived","0"))
                    recieveAudioPacketLost = int(stat.stats["audio_receive"].get("packetsLost","0"))
                    recieveAudioSsrc = stat.stats["audio_receive"].get("ssrc","0")
                    recieveAudioJitterReceived = int(stat.stats["audio_receive"].get("JitterReceived","0"))
                
                sentAudioPacketLost = 0
                sentAudioPacketSent =0
                sentAudioSsrc = ""
                sentAudioJitterReceived = 0
                sentAudioRtt = ""

                if "audio_sent" in stat.stats:    
                    sentAudioPacketLost = int(stat.stats["audio_sent"].get("packetsLost","0"))
                    sentAudioPacketSent = int(stat.stats["audio_sent"].get("packetsSent","0"))
                    sentAudioSsrc = stat.stats["audio_sent"].get("ssrc","0")
                    sentAudioCodecBitrate = int(stat.stats["audio_sent"].get("CodecBitrate","0"))
                    sentAudioJitterReceived = int(stat.stats["audio_sent"].get("JitterReceived","0"))
                    sentAudioRtt = stat.stats["audio_sent"].get("Rtt","0")
                    
                activeRemoteAddress = stat.stats.get("active_remote_address","N/A")

                detailsSheet.cell(column=3, row=detailRow, value=mos)
                jitter_cell = detailsSheet.cell(column=4, row=detailRow, value=jiiter)
                if int(jiiter) > jitter_thrashhold:
                    jitter_cell.fill = PatternFill("solid", start_color="FFA500")

                delayCell = detailsSheet.cell(column=5, row=detailRow, value=delay)

                if int(delay) > delay_thrashold:
                    delayCell.fill = PatternFill("solid", start_color="FFA500")

                packet_cell = detailsSheet.cell(column=6, row=detailRow, value=packet_loss)
                if float(packet_loss) > packloss_thrashold:
                    # print("Hit packet_loss  Thrashhold")
                    packet_cell.fill = PatternFill("solid", start_color="FFA500")
                    
                detailsSheet.cell(column=7, row=detailRow, value=r_value)
                detailsSheet.cell(column=8, row=detailRow, value=recieveAudioPacketsReceived)
                detailsSheet.cell(column=9, row=detailRow, value=recieveAudioPacketLost)
                detailsSheet.cell(column=10, row=detailRow, value=recieveAudioSsrc)
                detailsSheet.cell(column=11, row=detailRow, value=recieveAudioJitterReceived)
                if sentAudioPacketLost != None:
                    detailsSheet.cell(column=12, row=detailRow, value=sentAudioPacketSent)
                    detailsSheet.cell(column=13, row=detailRow, value=sentAudioPacketLost)
                    detailsSheet.cell(column=14, row=detailRow, value=sentAudioSsrc)
                    detailsSheet.cell(column=15, row=detailRow, value=sentAudioCodecBitrate)
                    detailsSheet.cell(column=16, row=detailRow, value=sentAudioJitterReceived)
                    detailsSheet.cell(column=17, row=detailRow, value=sentAudioRtt)
                
                detailsSheet.cell(column=18, row=detailRow, value=activeRemoteAddress)
                if stat.isCallHold == True:
                    for cell in detailsSheet[detailRow]:
                        cell.fill = PatternFill("solid", start_color="6495AD")

                detailRow += 1
                
                    

                self.adjustDemnetions(detailsSheet)

            row += 1
        self.adjustDemnetions(sheet)
        try:
            os.remove(self.fileName)
        except FileNotFoundError:
            pass

        mycell = sheet['B2']
        sheet.freeze_panes = mycell 
        workbook.save(self.fileName)
    
    def adjustDemnetions(self,sheet):
         for column in sheet.iter_cols():
            name = get_column_letter(column[0].column)
            new_col_length = max(len(str(cell.value)) for cell in column)
            sheet.column_dimensions[name].width = new_col_length+2 # Added a extra bit for padding

class Suspicious:
    def __init__(self,reason):
        self.reason = reason